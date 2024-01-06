import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# import win32com.client

# excel_app = win32com.client.Dispatch("Excel.Application")
# excel_app.Visible = False

# sys.exit()

def main():
    io_file = 'sim.xcfg'
    src_file_path = 'sim.xlsx'
    out_row_id = 3

    # root_dir = ''
    root_dir = '.'

    if root_dir:
        src_file_path = os.path.join(root_dir, src_file_path)

    src_file_path = os.path.abspath(src_file_path)

    with open(io_file, 'r') as fid:
        io_str = fid.readlines()

    io_data = [k.strip().split('\t') for k in io_str if k.strip() and not k.startswith('#')]
    src_book = load_workbook(src_file_path)

    prev_datum = None
    for _datum in io_data:

        if prev_datum is not None:
            _datum = [prev_datum[i] if k == '__' else k for i, k in enumerate(_datum)]

        src_sheet_name, src_start_col, src_start_row, src_end_col, src_end_row, \
        dst_sheet_name, dst_start_col, dst_start_row, dst_end_col, dst_end_row, \
        src_row_diff, dst_col_diff, n_reps = _datum

        src_row_diff = int(src_row_diff)
        dst_col_diff = int(dst_col_diff)
        n_reps = int(n_reps)

        # src_data = pd.read_excel(src_file_path, sheet_name=src_sheet)
        # src_df = pd.DataFrame(src_data)
        dst_sheet_id = src_book.sheetnames.index(dst_sheet_name)
        dst_sheet = src_book.worksheets[dst_sheet_id]

        src_sheet_id = src_book.sheetnames.index(src_sheet_name)
        src_sheet = src_book.worksheets[src_sheet_id]

        src_start_col_num = ord(src_start_col)
        src_end_col_num = ord(src_end_col)

        dst_start_col_num = ord(dst_start_col)
        dst_end_col_num = ord(dst_end_col)

        src_start_row, src_end_row = int(src_start_row), int(src_end_row)
        dst_start_row, dst_end_row = int(dst_start_row), int(dst_end_row)

        for rep_id in range(n_reps):

            src_cols = list(range(src_start_col_num, src_end_col_num + 1))
            dst_cols = list(range(dst_start_col_num, dst_end_col_num + 1))

            src_rows = list(range(src_start_row, src_end_row + 1))
            dst_rows = list(range(dst_start_row, dst_end_row + 1))

            assert len(src_cols) == len(dst_cols), "mismatch between src_cols and dst_cols len"
            assert len(src_rows) == len(dst_rows), "mismatch between src_rows and dst_rows len"

            for col_id, src_col_num in enumerate(src_cols):
                src_col_str = chr(src_col_num)
                dst_col_num = dst_cols[col_id]
                dst_col_str = chr(dst_col_num)

                src_col_idx = column_index_from_string(src_col_str)
                dst_col_idx = column_index_from_string(dst_col_str)

                for row_id, src_row_num in enumerate(src_rows):
                    dst_row_num = dst_rows[row_id]

                    # src_cell_id = '{}{}'.format(src_col_str, src_row_num)
                    # dst_cell_id = '{}{}'.format(dst_col_str, dst_row_num)

                    src_cell = src_sheet.cell(row=src_row_num, column=src_col_idx)
                    src_val = src_cell.value
                    dst_cell = dst_sheet.cell(row=dst_row_num, column=dst_col_idx)

                    dst_cell.value = src_val

            src_start_row += src_row_diff
            src_end_row += src_row_diff

            n_dst_cols = dst_end_col_num - dst_start_col_num
            dst_start_col_num += n_dst_cols + dst_col_diff
            dst_end_col_num = dst_start_col_num + n_dst_cols

        prev_datum = _datum

    # src_file_name = os.path.splitext(os.path.basename(src_file_path))[0]
    # src_file_dir = os.path.dirname(src_file_path)
    # dst_file_name = src_file_name + '_proc.xlsx'
    # dst_file_path = os.path.join(src_file_dir, dst_file_name)

    dst_file_path = src_file_path

    src_book.save(dst_file_path)


if __name__ == '__main__':
    main()

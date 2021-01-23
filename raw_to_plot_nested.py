import os
import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from ast import literal_eval
import paramparse


# sys.exit()

def run_op(op, src_book, local_offsets, offsets):
    assert isinstance(op, (str,)), 'Invalid op: {}'.format(op)

    src_sheet_name, src_start_col, src_start_row, src_end_col, src_end_row, \
    dst_sheet_name, dst_start_col, dst_start_row, dst_end_col, dst_end_row, = op.split(':')

    dst_sheet_id = src_book.sheetnames.index(dst_sheet_name)
    dst_sheet = src_book.worksheets[dst_sheet_id]

    src_sheet_id = src_book.sheetnames.index(src_sheet_name)
    src_sheet = src_book.worksheets[src_sheet_id]

    src_start_col_idx = column_index_from_string(src_start_col)
    src_end_col_idx = column_index_from_string(src_end_col)

    dst_start_col_idx = column_index_from_string(dst_start_col)
    dst_end_col_idx = column_index_from_string(dst_end_col)

    src_start_row, src_end_row = int(src_start_row), int(src_end_row)
    dst_start_row, dst_end_row = int(dst_start_row), int(dst_end_row)

    src_cols = list(range(src_start_col_idx, src_end_col_idx + 1))
    dst_cols = list(range(dst_start_col_idx, dst_end_col_idx + 1))

    src_rows = list(range(src_start_row, src_end_row + 1))
    dst_rows = list(range(dst_start_row, dst_end_row + 1))

    n_src_cols = len(src_cols)
    n_dst_cols = len(dst_cols)
    n_src_rows = len(src_rows)
    n_dst_rows = len(dst_rows)

    if n_src_cols == n_dst_rows and n_src_rows == n_dst_cols and n_src_cols != n_src_rows:
        transposed_dst = 1
        print('using transposed destination')
    else:
        assert n_src_cols == n_dst_cols, "mismatch between src_cols and dst_cols len"
        assert n_src_rows == n_dst_rows, "mismatch between src_rows and dst_rows len"

        transposed_dst = 0

    src_rows = [k + local_offsets[0] + offsets[0] for k in src_rows]
    src_cols = [k + local_offsets[1] + offsets[1] for k in src_cols]
    dst_rows = [k + local_offsets[2] + offsets[2] for k in dst_rows]
    dst_cols = [k + local_offsets[3] + offsets[3] for k in dst_cols]

    print(f'{src_sheet_name}::'
          f'{get_column_letter(min(src_cols))}{min(src_rows)}:'
          f'{get_column_letter(max(src_cols))}{max(src_rows)} --> '
          f'{dst_sheet_name}::'
          f'{get_column_letter(min(dst_cols))}{min(dst_rows)}:'
          f'{get_column_letter(max(dst_cols))}{max(dst_rows)}'
          )

    # return

    for src_col_id, src_col_num in enumerate(src_cols):
        for src_row_id, src_row_num in enumerate(src_rows):
            if not transposed_dst:
                dst_col_num = dst_cols[src_col_id]
                dst_row_num = dst_rows[src_row_id]
            else:
                dst_col_num = dst_cols[src_row_id]
                dst_row_num = dst_rows[src_col_id]

            # src_cell_id = '{}{}'.format(src_col_str, src_row_num)
            # dst_cell_id = '{}{}'.format(dst_col_str, dst_row_num)

            src_cell = src_sheet.cell(row=src_row_num, column=src_col_num)
            src_val = src_cell.value
            dst_cell = dst_sheet.cell(row=dst_row_num, column=dst_col_num)

            dst_cell.value = src_val


from operator import add


def process_io_dict(io_dict, src_book, offsets=(0, 0, 0, 0)):
    assert isinstance(io_dict, (dict,)), 'Invalid io_dict: {}'.format(io_dict)

    reps = io_dict['reps']
    ops = io_dict['ops']

    assert isinstance(ops, (list, str)), 'Invalid ops: {}'.format(ops)

    src_row, src_col, dst_row, dst_col, n_reps = map(int, reps.split(':'))

    local_offsets = (0, 0, 0, 0)

    for rep_id in range(n_reps):
        if isinstance(ops, str):
            ops = [ops, ]

        for op in ops:
            if isinstance(op, dict):
                _offsets = tuple(map(add, offsets, local_offsets))
                process_io_dict(op, src_book, _offsets)
            else:
                run_op(op, src_book, local_offsets, offsets)

        local_offsets = tuple(map(add, local_offsets, (src_row, src_col, dst_row, dst_col)))
        print('\n')
    print('\n\n')


class RTPNParams:
    in_place = 1
    # io_file = 'fence_gen1.xcfg'
    io_file = 'fence_gen1_circ.xcfg'
    # io_file = 'fence_gen1_cmb.xcfg'
    src_file_path = 'fence_gen1.xlsx'
    root_dir = 'C:/UofA/Acamp/fence_code/scripts/scripts_gen1/fence_results'


def main():
    params = RTPNParams()
    paramparse.process(params)

    root_dir = params.root_dir
    io_file = params.io_file
    src_file_path = params.src_file_path

    if root_dir:
        src_file_path = os.path.join(root_dir, src_file_path)
        io_file = os.path.join(root_dir, io_file)

    src_file_path = os.path.abspath(src_file_path)

    with open(io_file, 'r') as fid:
        io_lines = fid.readlines()

        io_lines = [k.strip() for k in io_lines if k.strip() and not k.startswith('#')]
        io_str = '\n'.join(io_lines)
        io_str = '[{}]'.format(io_str.strip())

        ops = literal_eval(io_str)
        # io_dict = json.load(fid)

    src_book = load_workbook(src_file_path)

    for op in ops:
        process_io_dict(op, src_book)

    if not params.in_place:
        src_file_name = os.path.splitext(os.path.basename(src_file_path))[0]
        src_file_dir = os.path.dirname(src_file_path)
        dst_file_name = src_file_name + '_proc.xlsx'
        dst_file_path = os.path.join(src_file_dir, dst_file_name)

        src_book.save(dst_file_path)

    else:

        dst_file_path = src_file_path

    import win32com.client

    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False

    try:
        src_book.save(dst_file_path)
    except PermissionError:
        excel_app.Workbooks.Open(dst_file_path).Close()
        # wb.Close()
        src_book.save(dst_file_path)

    excel_app.Workbooks.Open(dst_file_path)


if __name__ == '__main__':
    main()
